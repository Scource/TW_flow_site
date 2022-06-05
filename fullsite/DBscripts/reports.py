from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from django.http import HttpResponse
from django.core.files.base import ContentFile

from .models import Report, ReportItem
from .scripts import choose_from, generate_WIRE_objects, update_WIRE_dgmb_objects_list

reports_name_dict={
    'dbs_peak_values':'Dane zawyzone i zera',
    'dbs_statuses': 'Raport statusow',
    'dbs_check_node': 'Raport wezlow',
    'dbs_check_conf': 'Raport Konf i Zatw',
    'dbs_get_PPE': 'Zestawienie PPE',
    'dbs_check_MB_PPE': 'Porownanie MBvsPPE',
    'dbs_RDG': 'RDG',
}

def create_report(user, url, kwargs):
    main_rap = Report(base_number=1, report_author=user,
                      report_type=set_report_type(url))
    main_rap.save()
    if url in ('dbs_get_PPE', 'dbs_RDG'):
        for pob in kwargs['POB_elements']:
            if pob.cspr_id is not None:
                if url == 'dbs_RDG':
                    create_report_item(user, main_rap, url, kwargs, pob, strow=6)
                else:
                    create_report_item(user, main_rap, url, kwargs, pob)

    else:
        create_report_item(user, main_rap, url, kwargs)

    main_rap.end = datetime.now()
    main_rap.save()
    return main_rap.id

def set_report_name(url, kwargs, pob=None):
    if 'date_to' in kwargs:
        date_to_str = f"_{kwargs.get('date_to')}"
    else:
        date_to_str=""
    if pob:
        pob_str = f'_{pob.code}'
    else:
        pob_str=""
    return f"{reports_name_dict.get(url)}{pob_str}_{kwargs.get('date_from')}{date_to_str}.xlsx"

def set_report_type(url):
    return f"{reports_name_dict.get(url)}"


def create_report_item(user, main_rap, url, kwargs, pob=None, strow=0):
    in_memory_fp = BytesIO()
    doc = ReportItem(
        report=main_rap, report_name=set_report_name(url, kwargs, pob))
    doc.save()
    function = choose_from(url, 2)
    data = function(user, kwargs, pob) 
    writer = pd.ExcelWriter(in_memory_fp, engine='xlsxwriter')
    data.to_excel(writer, sheet_name='Raport', startrow=strow)
    writer.save()
    xlsx_data = ContentFile(in_memory_fp.getvalue())
    doc.report_file.save(doc.report_name, xlsx_data)
    doc.end = datetime.now()
    doc.save()
    if url=='dbs_RDG':
        add_head_data(doc.report_file, kwargs, pob)
    if url=='dbs_check_MB_PPE':
        color_diffs(doc.report_file)


def color_diffs(file):
    wb = load_workbook(filename=file.path)
    sheetname = wb.get_sheet_by_name('Raport')
    for col in sheetname.iter_cols(max_col=1):
        for cell in col:
            if cell.value=='ROZNICA': 
                wiersz=cell.row
                for c in sheetname[cell.row]:
                    if c.value!='ROZNICA' and c.value is not None:
                        if c.value>1 or c.value<-1:
                            c.fill = PatternFill(fgColor='FF0000', fill_type = 'solid')
    wb.save(file.path)

def add_head_data(file, kwargs, pob):
    od = datetime.strftime(kwargs.get('date_from'), "%d-%m-%Y,%H:%M")
    do = datetime.strftime((kwargs.get('date_to') + timedelta(days=1)), "%d-%m-%Y,%H:%M")

    date=f'{od} - {do}'
    name_str=pob.code

    wb = load_workbook(filename=file.path)
    sheetname = wb.get_sheet_by_name('Raport')
    sheetname['B3'] = str('Dla formuły: ')
    sheetname['B4'] = str('Za okres:')
    sheetname['D3'] = str(f'{name_str} [ kW ]')
    sheetname['D4'] = str(date)
    wb.save(file.path)

def create_for_WIRE(user, data):
    #select objects for correction
    print(data['get_all_FBT'])
    if data['get_all_FBT']:
        list_with_changes=generate_WIRE_objects(user, data)
        data['POB_elements']=list_with_changes
    
    #update WIRE db
    update_WIRE_dgmb_objects_list(data['POB_elements'])

    response=create_bat_file(user, data)

    #if needed create RDG for all MB with differences
    if data['rdg']:
        create_report(user, url='dbs_RDG', kwargs=data)
    return response

def create_bat_file(user, data):
    #create string out of wire ids and add EPS OO and OP
    ppom_list = set(data['POB_elements'].values_list('wire_id', flat=True))
    ppom_list.add('508')
    ppom_list.add('509')
    ppom_list_string=', '.join(str(e) for e in ppom_list if e is not None)

    #set correction date fe: -M2, M4 etc
    corr_date=dif_months(data['date_from'])

    #serve file ass response
    filename = f"{user.username}_{corr_date}_{data['date_from']}-{data['date_to']}.bat"
    content = (f"""c:\innsoft\wire\import\wibdskome.exe "{corr_date}" "{ppom_list_string}" "-T" && echo 
    C:\innsoft\Wire\Import\webddgmb.exe wire wire innsoft {user.wire_user} 1191 -Q {corr_date}""")
    response = HttpResponse(content, content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
    return response

def dif_months(cor):
    now=datetime.now()
    return f"-M{(now.year - cor.year)*12 + now.month - cor.month}"